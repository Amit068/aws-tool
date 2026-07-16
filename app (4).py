import streamlit as st
import pdfplumber
import re
import hashlib
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
from io import BytesIO
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────
# Page Config
# ─────────────────────────────────────────────
st.set_page_config(
    page_title="AWS Invoice Dashboard",
    page_icon="https://upload.wikimedia.org/wikipedia/commons/9/93/Amazon_Web_Services_Logo.svg",
    layout="wide",
)

# ─────────────────────────────────────────────
# Custom CSS
# ─────────────────────────────────────────────
st.markdown("""
<style>
    .stApp { background-color: #EBF5FB !important; }
    .main  { background-color: #EBF5FB !important; }
    section[data-testid="stSidebar"] { background-color: #BBDEFB !important; }
    .block-container { padding-top: 0 !important; max-width: 1400px; }
    #MainMenu, header, footer,
    [data-testid="stToolbar"],
    [data-testid="stDecoration"],
    [data-testid="stStatusWidget"],
    .stDeployButton {
        display: none !important; visibility: hidden !important; height: 0 !important;
    }
    .aws-hero {
        background: #ffffff; border-bottom: 4px solid #FF9900;
        padding: 24px 40px 22px 40px; margin: 0 0 1.6rem 0;
        display: flex; align-items: center; gap: 24px;
        border-radius: 0 0 16px 16px;
        box-shadow: 0 4px 20px rgba(21,101,192,0.16);
    }
    .aws-logo-wrap {
        background: #ffffff; border-radius: 12px; padding: 14px 18px 8px 18px;
        display: flex; align-items: center; justify-content: center;
        flex-shrink: 0; width: 160px; height: 74px; margin-top: 0;
        box-shadow: 0 6px 18px rgba(13,43,78,0.12);
    }
    .aws-logo-wrap img { width: 130px; height: auto; display: block;
                         object-fit: contain; transform: translateY(6px); }
    .hero-text h1 { color: #0D2B4E; font-size: 30px; font-weight: 700;
                    margin: 0 0 4px 0; line-height: 1.15; }
    .hero-text p  { color: #1565C0; font-size: 14px; margin: 0; }
    .hero-badge {
        margin-left: auto; background: #FFF8EC;
        border: 1px solid rgba(255,153,0,0.6); color: #FF9900;
        padding: 6px 18px; border-radius: 20px;
        font-size: 13px; font-weight: 700; white-space: nowrap;
    }
    .mode-card {
        background: #ffffff; border-radius: 14px; padding: 22px 24px;
        border: 2.5px solid #90CAF9; cursor: pointer;
        box-shadow: 0 2px 10px rgba(25,118,210,0.08);
        transition: all 0.18s; margin-bottom: 4px;
    }
    .mode-card:hover { border-color: #FF9900; box-shadow: 0 4px 20px rgba(255,153,0,0.15); }
    .mode-card.active { border-color: #1565C0; border-width: 3px;
                        box-shadow: 0 4px 18px rgba(21,101,192,0.2); }
    .mode-title { font-size: 16px; font-weight: 700; color: #0D2B4E; margin-bottom: 6px; }
    .mode-desc  { font-size: 12px; color: #1565C0; line-height: 1.5; }
    .metric-card {
        background: #ffffff; border-radius: 12px; padding: 18px 20px;
        border: 1px solid #BBDEFB; border-top: 4px solid #1976D2;
        box-shadow: 0 2px 8px rgba(25,118,210,0.08);
    }
    .metric-label { font-size: 11px; color: #1565C0; font-weight: 700;
                    letter-spacing: 0.8px; text-transform: uppercase; }
    .metric-value { font-size: 26px; color: #0D2B4E; font-weight: 700; margin-top: 6px; }
    .metric-sub   { font-size: 11px; color: #64B5F6; margin-top: 3px; }
    .section-title {
        font-size: 15px; font-weight: 700; color: #1565C0;
        border-bottom: 2px solid #90CAF9;
        padding-bottom: 8px; margin: 24px 0 16px 0;
        text-transform: uppercase; letter-spacing: 0.5px;
    }
    .insight-box {
        background: #E3F2FD; border-radius: 10px; padding: 14px 18px;
        border-left: 4px solid #1976D2; margin: 8px 0; color: #0D2B4E;
    }
    .insight-box b { color: #1565C0; }
    div[data-testid="stFileUploader"] {
        border: 1.5px dashed #64B5F6 !important; border-radius: 12px !important;
        padding: 14px !important; background: #ffffff !important;
        box-shadow: 0 4px 16px rgba(25,118,210,0.08) !important;
    }
    div[data-testid="stFileUploader"]:hover {
        border-color: #FF9900 !important;
        box-shadow: 0 6px 20px rgba(255,153,0,0.12) !important;
    }
    div[data-testid="stButton"] button {
        background: linear-gradient(135deg, #FF9900, #E68A00) !important;
        color: #ffffff !important; font-weight: 700 !important;
        border: none !important; border-radius: 10px !important;
        min-height: 42px !important;
        box-shadow: 0 8px 18px rgba(255,153,0,0.22) !important;
    }
    div[data-testid="stButton"] button:disabled {
        background: #D6E4F0 !important; color: #6B7F93 !important;
        box-shadow: none !important;
    }
    div[data-testid="stDownloadButton"] button {
        background: linear-gradient(135deg, #FF9900, #E68A00) !important;
        color: #ffffff !important; font-weight: 700 !important;
        border: none !important; border-radius: 10px !important;
        font-size: 16px !important; padding: 14px !important;
    }
    button[data-baseweb="tab"] { color: #1976D2 !important; font-weight: 600 !important; }
    button[data-baseweb="tab"][aria-selected="true"] {
        color: #1565C0 !important; border-bottom-color: #FF9900 !important;
    }
    hr { border-color: #90CAF9 !important; }
    .chart-box { background: transparent; border-radius: 14px;
                 padding: 0; border: 0; box-shadow: none; margin-bottom: 0; }
    .payer-card {
        background: linear-gradient(135deg, #E3F2FD, #BBDEFB);
        border-radius: 12px; padding: 16px 20px;
        border-left: 5px solid #1565C0; margin-bottom: 20px;
        display: flex; align-items: center; gap: 20px;
    }
    p, span, label { color: #0D2B4E; }
    h1, h2, h3, h4 { color: #0D2B4E !important; }
    ::-webkit-scrollbar { width: 6px; height: 6px; }
    ::-webkit-scrollbar-track { background: #EBF5FB; }
    ::-webkit-scrollbar-thumb { background: #64B5F6; border-radius: 3px; }
</style>
""", unsafe_allow_html=True)


# ─────────────────────────────────────────────
# HERO HEADER
# ─────────────────────────────────────────────
st.markdown("""
<div class="aws-hero">
    <div class="aws-logo-wrap">
        <img src="https://upload.wikimedia.org/wikipedia/commons/9/93/Amazon_Web_Services_Logo.svg"
             alt="AWS Logo"/>
    </div>
    <div class="hero-text">
        <h1 style="color:#0D2B4E;font-size:32px;font-weight:700;margin:0 0 4px 0;">Invoice Dashboard</h1>
        <p style="color:#1565C0;font-size:14px;margin:0;">
            Upload PDF &rarr; Extract Data &rarr; View Charts &amp; Insights
        </p>
    </div>
    <div class="hero-badge">&#9679;&nbsp;AWS Cost Analytics</div>
</div>
""", unsafe_allow_html=True)

# Session data warning
st.markdown("""
<div style="background:#FFFBF2;border:1px solid #FFD58A;border-left:4px solid #FF9900;
    border-radius:12px;padding:14px 18px;margin-bottom:18px;
    display:flex;align-items:center;gap:12px;box-shadow:0 4px 14px rgba(255,153,0,0.08);">
    <span style="font-size:20px;">⚠️</span>
    <div>
        <span style="color:#FF9900;font-weight:700;font-size:14px;">Important — Session Data Warning</span><br>
        <span style="color:#0D2B4E;font-size:13px;">
            Refreshing or closing this tab will <b style="color:#FF9900;">permanently clear all uploaded data.</b>
            Please download your Excel report before leaving.
        </span>
    </div>
</div>
""", unsafe_allow_html=True)


# ═══════════════════════════════════════════════════════════
# MODE SELECTION
# ═══════════════════════════════════════════════════════════
st.markdown('<div class="section-title">🔀 Select Report Type</div>', unsafe_allow_html=True)

col_m1, col_m2 = st.columns(2)

with col_m1:
    st.markdown("""
    <div class="mode-card">
        <div class="mode-title">📄 Mode 1 — Account-wise Invoice</div>
        <div class="mode-desc">
            Upload your AWS Consolidated Invoice PDF.<br>
            Get per-account charges, discounts, GST & net totals at a glance.<br>
            <b>PDF source:</b> AWS Billing → Bills → Download PDF Invoice
        </div>
    </div>
    """, unsafe_allow_html=True)

with col_m2:
    st.markdown("""
    <div class="mode-card">
        <div class="mode-title">📊 Mode 2 — Service-wise Cost per Account</div>
        <div class="mode-desc">
            Upload the same AWS Consolidated Invoice PDF.<br>
            Get an Account &times; Service table — Net Cost = Charges − Savings Plan.<br>
            <b>PDF source:</b> AWS Billing → Bills → Download PDF Invoice
        </div>
    </div>
    """, unsafe_allow_html=True)

selected_mode = st.radio(
    "Select mode:",
    ["📄 Mode 1 — Account-wise Invoice (Consolidated PDF)",
     "📊 Mode 2 — Service-wise Cost Breakdown (Bill PDF)"],
    horizontal=True,
    label_visibility="collapsed",
)
mode = 1 if "Mode 1" in selected_mode else 2

st.markdown("---")


# ═══════════════════════════════════════════════════════════════════════════════
# ── COMMON PARSE FUNCTIONS ────────────────────────────────────────────────────
# ═══════════════════════════════════════════════════════════════════════════════

# ── Mode 1 helpers ────────────────────────────────────────────────────────────
_SAVINGS_SUB = re.compile(r'^Savings Plan \(Charges covered by Savings Plans\)')
_SUB_STARTS  = ('Charges ', 'Tax ', 'GST ', 'Credits ', 'Discount (')

def _usd(text):
    m = re.search(r'-?USD\s*([\d,]+\.?\d*)', text)
    if not m:
        return 0.0
    val = float(m.group(1).replace(',', ''))
    return -val if re.search(r'(?<!\w)-USD', text) or text.strip().startswith('-') else val

def _is_sub(line):
    if _SAVINGS_SUB.match(line):
        return True
    return any(line.startswith(p) for p in _SUB_STARTS)

def _pdf_hash(pdf_bytes: bytes) -> str:
    return hashlib.md5(pdf_bytes).hexdigest()

@st.cache_data(show_spinner=False, max_entries=20)
def _extract_text_invoice(pdf_hash: str, pdf_bytes: bytes) -> str:
    """Extract full text from invoice PDF — cached by MD5 hash."""
    with pdfplumber.open(BytesIO(pdf_bytes)) as pdf:
        pages_text = []
        for page in pdf.pages:
            t = page.extract_text(x_tolerance=2, y_tolerance=2)
            if t:
                pages_text.append(t)
    return "\n".join(pages_text)

@st.cache_data(show_spinner=False, max_entries=20)
def parse_invoice_pdf(pdf_hash: str, pdf_bytes: bytes):
    account_rows = []
    service_data = {}

    full_text = _extract_text_invoice(pdf_hash, pdf_bytes)

    payer_account = None
    payer_name    = None
    acct_m = re.search(r'Account number[:\s]*\n?\s*(\d{12})', full_text)
    if acct_m:
        payer_account = acct_m.group(1).strip()
    bill_m = re.search(r'Bill to Address[:\s]*\n\s*(.+)', full_text)
    if bill_m:
        payer_name = bill_m.group(1).strip()

    blocks = re.split(r'Summary for Linked Account\s*\n', full_text)

    for block in blocks[1:]:
        lines = [l.strip() for l in block.split('\n') if l.strip()]
        if not lines:
            continue

        hm = re.match(r'^(.+?)\s*\((\d{12})\)\s+USD\s+([\d,]+\.?\d*)', lines[0])
        if not hm:
            continue

        acc_name = hm.group(1).strip()
        acc_id   = hm.group(2).strip()

        charges = savings_plan = private_rate = bundled_dis = 0.0
        edp_dis = credits = gst = total_alloc = 0.0
        in_summary = True
        in_detail  = False
        svc_name   = None
        svc_charges = svc_savings = 0.0
        cur_svcs   = {}

        for line in lines[1:]:
            if 'total allocated for this statement' in line.lower():
                m = re.search(r'USD\s*([\d,]+\.?\d*)', line)
                total_alloc = float(m.group(1).replace(',', '')) if m else 0.0
                in_summary = False
                continue
            if 'Detail for Linked Account' in line:
                in_detail = True; in_summary = False
                continue
            if 'For line item details' in line or 'Account Activity Page' in line:
                break

            if in_summary:
                if re.match(r'^Charges\s+USD', line):
                    charges = abs(_usd(line))
                elif _SAVINGS_SUB.match(line):
                    savings_plan = abs(_usd(line))
                elif 'Private Rate Card' in line:
                    private_rate = abs(_usd(line))
                elif 'Bundled Discount' in line:
                    bundled_dis = abs(_usd(line))
                elif ('Distribution Program Discount' in line or
                      'Enterprise Discount Program' in line or
                      'AWS Distribution Program Discount' in line):
                    edp_dis = abs(_usd(line))
                elif re.match(r'^Credits\s+', line):
                    credits = _usd(line)
                elif re.match(r'^Tax\s+USD', line) or re.match(r'^GST\s+USD', line):
                    gst += abs(_usd(line))
            elif in_detail:
                is_sub    = _is_sub(line)
                svc_match = re.match(r'^(.+?)\s+USD\s+([\d,]+\.?\d*)$', line)
                if svc_match and not is_sub:
                    if svc_name:
                        net = svc_charges - svc_savings
                        cur_svcs[svc_name] = cur_svcs.get(svc_name, 0.0) + net
                    svc_name    = svc_match.group(1).strip()
                    svc_charges = svc_savings = 0.0
                elif svc_name:
                    if re.match(r'^Charges\s+USD', line):
                        svc_charges = abs(_usd(line))
                    elif _SAVINGS_SUB.match(line):
                        svc_savings = abs(_usd(line))

        if svc_name and in_detail:
            cur_svcs[svc_name] = cur_svcs.get(svc_name, 0.0) + (svc_charges - svc_savings)

        account_rows.append({
            'Account ID'                 : acc_id,
            'Account Name'               : acc_name,
            'Charges (USD)'              : round(charges, 2),
            'Savings Plan (USD)'         : round(-savings_plan, 2),
            'Private Rate Card (USD)'    : round(-private_rate, 2),
            'Bundled Discount (USD)'     : round(-bundled_dis, 2),
            'EDP / Dist. Discount (USD)' : round(-edp_dis, 2),
            'Credits (USD)'              : round(credits, 2),
            'GST (USD)'                  : round(gst, 2),
            'Total (USD)'                : round(total_alloc, 2),
        })

        if acc_id not in service_data:
            service_data[acc_id] = {'_name': acc_name}
        for svc, val in cur_svcs.items():
            service_data[acc_id][svc] = round(service_data[acc_id].get(svc, 0.0) + val, 2)

    return account_rows, service_data, payer_account, payer_name


# ── Mode 2 helpers ────────────────────────────────────────────────────────────
# Mode 2 re-uses the same proven "Summary for Linked Account" / "Detail for
# Linked Account" parser as Mode 1 (parse_invoice_pdf above), because that is
# the actual structure of the AWS Statement PDF. It already computes, per
# account, per service: Net Cost = Charges − Savings Plan (if a Savings Plan
# line is present for that service; otherwise Net Cost = Charges as-is).
# Mode 2 simply pivots that data into an Account x Service table.

def build_service_pivot_df(service_data: dict) -> pd.DataFrame:
    """Account ID (rows) x Service (columns) = Net Cost (Charges - Savings Plan)."""
    all_services = sorted({s for d in service_data.values() for s in d if s != '_name'})
    rows = []
    for acc_id, acc_data in service_data.items():
        row = {'Account ID': acc_id}
        for svc in all_services:
            row[svc] = round(acc_data.get(svc, 0.0) or 0.0, 2)
        rows.append(row)
    df = pd.DataFrame(rows)
    return df


# ═══════════════════════════════════════════════════════════════════════════════
# ── EXCEL BUILDERS ────────────────────────────────────────────────────────────
# ═══════════════════════════════════════════════════════════════════════════════
HDR_BG = "1F4E79"; HDR_FG = "FFFFFF"
ALT_BG = "D6E4F0"; NEG_FG = "C00000"
TOTAL_BG = "BDD7EE"; TOTAL_FG = "1F4E79"

def _border():
    s = Side(style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def _hdr(cell, value):
    cell.value = value
    cell.font  = Font(bold=True, color=HDR_FG, name="Arial", size=10)
    cell.fill  = PatternFill("solid", start_color=HDR_BG)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = _border()

def _dat(cell, value, row_idx, align="left", neg=False):
    cell.value = value
    bg = ALT_BG if row_idx % 2 == 0 else "FFFFFF"
    cell.fill  = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal=align, vertical="center")
    color = NEG_FG if neg and isinstance(value, (int, float)) and value < 0 else "000000"
    cell.font = Font(name="Arial", size=10, color=color)
    cell.border = _border()

def _total_cell(cell, value, align="right"):
    cell.value = value
    cell.font  = Font(bold=True, color=TOTAL_FG, name="Arial", size=10)
    cell.fill  = PatternFill("solid", start_color=TOTAL_BG)
    cell.alignment = Alignment(horizontal=align, vertical="center")
    cell.border = _border()

def _auto_width(ws, min_w=10, max_w=35):
    for col in ws.columns:
        length = max((len(str(c.value)) if c.value is not None else 0) for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max(min_w, min(length+3, max_w))

S1_HDRS = ['Account ID','Account Name','Charges (USD)','Savings Plan (USD)',
           'Private Rate Card (USD)','Bundled Discount (USD)',
           'EDP / Dist. Discount (USD)','Credits (USD)','GST (USD)','Total (USD)']
S1_NEG = {4,5,6,7,8}
S1_NUM = set(range(3, len(S1_HDRS)+1))

def build_invoice_excel(all_rows, all_svcs):
    wb  = Workbook()
    ws1 = wb.active
    ws1.title = "Account Details"
    ws1.row_dimensions[1].height = 32
    for ci, h in enumerate(S1_HDRS, 1):
        _hdr(ws1.cell(1, ci), h)
    for ri, row in enumerate(all_rows, 2):
        for ci, h in enumerate(S1_HDRS, 1):
            _dat(ws1.cell(ri, ci), row[h], ri,
                 align="right" if ci in S1_NUM else "left", neg=(ci in S1_NEG))
    tr = len(all_rows)+2
    _total_cell(ws1.cell(tr,1), "TOTAL", align="center")
    _total_cell(ws1.cell(tr,2), "")
    for ci, h in enumerate(S1_HDRS[2:], 3):
        col_vals = [r[h] for r in all_rows if isinstance(r[h], (int,float))]
        _total_cell(ws1.cell(tr,ci), round(sum(col_vals),2))
    _auto_width(ws1)
    ws1.freeze_panes = "A2"
    ws1.auto_filter.ref = ws1.dimensions

    ws2 = wb.create_sheet("Service Breakdown")
    ws2.row_dimensions[1].height = 48
    all_services = sorted({s for d in all_svcs.values() for s in d if s != '_name'})
    s2_hdrs = ['Account ID','Account Name'] + all_services
    for ci, h in enumerate(s2_hdrs, 1):
        _hdr(ws2.cell(1, ci), h)
    for ri, (acc_id, acc_data) in enumerate(all_svcs.items(), 2):
        _dat(ws2.cell(ri,1), acc_id, ri)
        _dat(ws2.cell(ri,2), acc_data.get('_name',''), ri)
        for ci, svc in enumerate(all_services, 3):
            val = acc_data.get(svc, None)
            _dat(ws2.cell(ri,ci), round(val,2) if val is not None else None, ri, align="right")
    tr2 = len(all_svcs)+2
    _total_cell(ws2.cell(tr2,1), "TOTAL", align="center")
    _total_cell(ws2.cell(tr2,2), "")
    for ci, svc in enumerate(all_services, 3):
        col_vals = [d.get(svc,0.0) for d in all_svcs.values() if isinstance(d.get(svc,0.0),(int,float))]
        _total_cell(ws2.cell(tr2,ci), round(sum(col_vals),2))
    _auto_width(ws2, min_w=12, max_w=30)
    ws2.freeze_panes = "A2"
    ws2.auto_filter.ref = ws2.dimensions

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def build_service_pivot_excel(df_pivot: pd.DataFrame):
    """Excel export for the Account x Service pivot (Net Cost = Charges - Savings Plan)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Service Breakdown"

    thin = Side(style='thin', color='BFBFBF')
    bdr  = Border(left=thin, right=thin, top=thin, bottom=thin)
    h_font = Font(name='Arial', bold=True, color='FFFFFF', size=10)
    h_fill = PatternFill('solid', start_color='1F4E79')
    d_font = Font(name='Arial', size=9)
    a_fill = PatternFill('solid', start_color='EBF3FB')
    w_fill = PatternFill('solid', start_color='FFFFFF')
    t_font = Font(name='Arial', bold=True, color='1F4E79', size=10)
    t_fill = PatternFill('solid', start_color='BDD7EE')
    num_fmt = '#,##0.00'

    headers = list(df_pivot.columns)  # 'Account ID' + service names
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = h_font; c.fill = h_fill
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = bdr
    ws.row_dimensions[1].height = 40
    ws.freeze_panes = 'B2'

    for ri, row in enumerate(df_pivot.itertuples(index=False), 2):
        fill = a_fill if ri % 2 == 0 else w_fill
        for ci, val in enumerate(row, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.font = d_font; c.fill = fill; c.border = bdr
            if ci == 1:
                c.alignment = Alignment(horizontal='left', vertical='center')
            else:
                c.number_format = num_fmt
                c.alignment = Alignment(horizontal='right', vertical='center')

    tr = len(df_pivot) + 2
    tc = ws.cell(row=tr, column=1, value="TOTAL")
    tc.font = t_font; tc.fill = t_fill; tc.border = bdr
    for ci, h in enumerate(headers[1:], 2):
        total = round(df_pivot[h].sum(), 2)
        c = ws.cell(row=tr, column=ci, value=total)
        c.font = t_font; c.fill = t_fill; c.border = bdr
        c.number_format = num_fmt; c.alignment = Alignment(horizontal='right')

    for col in ws.columns:
        length = max((len(str(c.value)) if c.value is not None else 0) for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max(12, min(length + 3, 32))
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{tr-1}"

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ═══════════════════════════════════════════════════════════════════════════════
# ══════════════════════  MODE 1 — ACCOUNT-WISE INVOICE  ═══════════════════════
# ═══════════════════════════════════════════════════════════════════════════════
if mode == 1:
    st.markdown('<div class="section-title">📂 Upload Consolidated Invoice PDF</div>',
                unsafe_allow_html=True)

    st.markdown("""
    <div style="background:#E3F2FD;border-radius:10px;padding:10px 14px;
                border-left:4px solid #1976D2;margin-bottom:12px;">
        <b style="color:#1565C0;">📄 Mode 1 — Consolidated Invoice PDF</b><br>
        <span style="color:#0D2B4E;font-size:12px;">
            Navigate to AWS Billing → Bills → Download PDF Invoice and upload it here.<br>
            <b>Only Consolidated Invoice PDFs are supported in this mode — do not upload Bill Summary or Activity PDFs.</b>
        </span>
    </div>
    """, unsafe_allow_html=True)

    invoice_files = st.file_uploader(
        "Upload Consolidated Invoice PDF(s)",
        type=["pdf"],
        accept_multiple_files=True,
        key="mode1_uploader",
        help="AWS Consolidated Invoice — contains Summary for Linked Accounts"
    )

    has_invoice = bool(invoice_files)

    if has_invoice:
        st.success(f"✅ {len(invoice_files)} Invoice PDF(s) ready to process")
    else:
        st.info("⬆️ Upload your Consolidated Invoice PDF(s) and click Submit.")

    uploaded_file_key_m1 = tuple((f.name, getattr(f,"size",0)) for f in (invoice_files or []))
    if st.session_state.get("m1_file_key") != uploaded_file_key_m1:
        st.session_state["m1_file_key"] = uploaded_file_key_m1
        st.session_state.pop("m1_parsed_data", None)
        # Clear excel caches for all payer keys
        for k in list(st.session_state.keys()):
            if k.startswith("m1_excel"):
                del st.session_state[k]

    submit_m1 = st.button("🚀 Submit & Process Invoice PDF",
                          type="primary", use_container_width=True,
                          disabled=not has_invoice)

    if not has_invoice:
        st.stop()

    if not submit_m1 and "m1_parsed_data" not in st.session_state:
        st.info("PDF selected. Click 'Submit' to extract and visualize data.")
        st.stop()

    # ── Parse ──
    if submit_m1:
        all_rows  = []
        all_svcs  = {}
        errors    = []
        payer_map = {}

        progress = st.progress(0, text="Reading Invoice PDFs...")
        for i, f in enumerate(invoice_files):
            try:
                pdf_bytes = f.getvalue()
                pdf_hash  = _pdf_hash(pdf_bytes)
                rows, svcs, p_acct, p_name = parse_invoice_pdf(pdf_hash, pdf_bytes)
                all_rows.extend(rows)
                for acc_id, acc_data in svcs.items():
                    if acc_id not in all_svcs:
                        all_svcs[acc_id] = acc_data.copy()
                    else:
                        for k, v in acc_data.items():
                            if k == '_name':
                                all_svcs[acc_id]['_name'] = v
                            else:
                                all_svcs[acc_id][k] = round(all_svcs[acc_id].get(k,0.0)+v, 2)
                pkey = p_acct or f"file_{i}"
                if pkey not in payer_map:
                    payer_map[pkey] = {
                        'account_id': p_acct or "Unknown",
                        'name': p_name or f.name.replace('.pdf',''),
                        'rows': [], 'svcs': {},
                    }
                payer_map[pkey]['rows'].extend(rows)
                for acc_id, acc_data in svcs.items():
                    pm = payer_map[pkey]['svcs']
                    if acc_id not in pm:
                        pm[acc_id] = acc_data.copy()
                    else:
                        for k, v in acc_data.items():
                            if k == '_name': pm[acc_id]['_name'] = v
                            else: pm[acc_id][k] = round(pm[acc_id].get(k,0.0)+v, 2)
            except Exception as e:
                errors.append(f"{f.name}: {e}")
            progress.progress(int((i+1)/len(invoice_files)*100), text=f"{f.name} done")

        progress.empty()
        st.session_state["m1_parsed_data"] = {
            "all_rows": all_rows, "all_svcs": all_svcs,
            "payer_map": payer_map, "errors": errors,
        }

    data = st.session_state.get("m1_parsed_data", {})
    all_rows  = data.get("all_rows", [])
    all_svcs  = data.get("all_svcs", {})
    payer_map = data.get("payer_map", {})
    errors    = data.get("errors", [])

    for e in errors:
        st.error(f"❌ {e}")

    if not all_rows:
        st.warning("⚠️ No data extracted. Please verify you have uploaded a valid Consolidated Invoice PDF.")
        st.stop()

    # ── Payer filter ──
    selected_payer_key  = None
    selected_payer_info = None
    if len(payer_map) > 1:
        payer_options = {"🌐 All Payers (Combined)": None}
        for k, v in payer_map.items():
            payer_options[f"💳 {v['name']} ({v['account_id']})"] = k
        selected_label = st.selectbox("Filter by Payer:", list(payer_options.keys()))
        selected_payer_key = payer_options[selected_label]
        if selected_payer_key:
            selected_payer_info = payer_map[selected_payer_key]

    display_rows = (selected_payer_info['rows'] if selected_payer_info else all_rows)
    display_svcs = (selected_payer_info['svcs'] if selected_payer_info else all_svcs)

    if not display_rows:
        st.warning("No data available for the selected payer.")
        st.stop()

    # DataFrames
    df_accounts = pd.DataFrame(display_rows)
    all_services = sorted({s for d in display_svcs.values() for s in d if s != '_name'})
    svc_rows = []
    for acc_id, acc_data in display_svcs.items():
        row = {'Account ID': acc_id, 'Account Name': acc_data.get('_name','')}
        for svc in all_services:
            row[svc] = acc_data.get(svc, 0.0) or 0.0
        svc_rows.append(row)
    df_services = pd.DataFrame(svc_rows)

    svc_totals = {}
    for acc_data in display_svcs.values():
        for k, v in acc_data.items():
            if k != '_name' and isinstance(v, (int,float)):
                svc_totals[k] = round(svc_totals.get(k,0.0)+v, 2)
    df_svc_totals = pd.DataFrame(
        sorted(svc_totals.items(), key=lambda x: x[1], reverse=True),
        columns=['Service','Total Cost (USD)']
    ).query("`Total Cost (USD)` > 0")

    # KPI Cards
    st.markdown('<div class="section-title">📊 Summary</div>', unsafe_allow_html=True)
    charges_before_savings = df_accounts['Charges (USD)'].sum()
    savings_plan_total = df_accounts.get('Savings Plan (USD)', pd.Series([0])).sum()
    total_charges = charges_before_savings + savings_plan_total
    total_bill    = df_accounts['Total (USD)'].sum()
    total_gst     = df_accounts['GST (USD)'].sum()
    total_credits = df_accounts.get('Credits (USD)', pd.Series([0])).sum()
    total_discount = (
        savings_plan_total +
        df_accounts.get('Private Rate Card (USD)', pd.Series([0])).sum() +
        df_accounts.get('Bundled Discount (USD)', pd.Series([0])).sum() +
        df_accounts.get('EDP / Dist. Discount (USD)', pd.Series([0])).sum()
    )
    top_service  = df_svc_totals.iloc[0]['Service'] if not df_svc_totals.empty else "N/A"
    top_svc_cost = df_svc_totals.iloc[0]['Total Cost (USD)'] if not df_svc_totals.empty else 0

    c1, c2, c3, c4, c5, c6 = st.columns(6)
    def kpi(col, icon, label, value, sub=""):
        col.markdown(f"""
        <div class="metric-card">
          <div class="metric-label">{icon} {label}</div>
          <div class="metric-value">{value}</div>
          <div class="metric-sub">{sub}</div>
        </div>""", unsafe_allow_html=True)

    total_pdfs = len(invoice_files or [])
    kpi(c1, "🏢", "Total Accounts",  str(len(display_rows)), f"{total_pdfs} PDF(s)")
    kpi(c2, "💰", "Total Charges",   f"${total_charges:,.2f}", "Charges - Savings Plan")
    kpi(c3, "🧾", "Total Bill",      f"${total_bill:,.2f}", "After all discounts")
    kpi(c4, "🏷️", "Total Discount",  f"${abs(total_discount):,.2f}", "SP + Private + Bundle + EDP")
    kpi(c5, "↩", "Invoice Credits",  f"${total_credits:,.2f}", "Credits from invoice")
    kpi(c6, "🔝", "Top Service",
        top_service[:16]+"…" if len(top_service)>16 else top_service,
        f"${top_svc_cost:,.2f}")

    st.markdown("<br>", unsafe_allow_html=True)

    # Charts
    chart_col1, chart_col2 = st.columns([1,1])

    with chart_col1:
        st.markdown("#### 🥧 Service Cost Distribution")
        pie_df = df_svc_totals.head(10)
        fig_pie = px.pie(pie_df, names='Service', values='Total Cost (USD)',
                         color_discrete_sequence=px.colors.sequential.Blues_r, hole=0.35)
        fig_pie.update_traces(textposition='inside', textinfo='percent+label',
                              textfont=dict(color='white', size=10),
                              marker=dict(line=dict(color='#ffffff', width=1.5)))
        fig_pie.update_layout(paper_bgcolor='#ffffff', plot_bgcolor='#ffffff',
                              font=dict(color='#0D2B4E', family='Arial'),
                              margin=dict(t=10,b=10,l=10,r=10),
                              legend=dict(orientation="v", x=1.02, y=0.5,
                                         bgcolor='rgba(0,0,0,0)', font=dict(color='#0D2B4E')),
                              height=380)
        st.plotly_chart(fig_pie, use_container_width=True)

    with chart_col2:
        st.markdown("#### 📊 Top Services by Cost")
        bar_df = df_svc_totals.head(12)
        fig_bar = px.bar(bar_df, x='Total Cost (USD)', y='Service', orientation='h',
                         color='Total Cost (USD)',
                         color_continuous_scale=[[0,'#87CEEB'],[0.5,'#29B6F6'],[1,'#FF9900']],
                         text='Total Cost (USD)')
        fig_bar.update_traces(texttemplate='$%{text:,.0f}', textposition='outside',
                              textfont=dict(color='#0D2B4E', size=11),
                              marker=dict(line=dict(color='#ffffff', width=0.5)))
        fig_bar.update_layout(paper_bgcolor='#ffffff', plot_bgcolor='#ffffff',
                              font=dict(color='#0D2B4E', family='Arial'),
                              yaxis=dict(autorange="reversed", color='#1565C0', gridcolor='#E3F2FD'),
                              xaxis=dict(color='#1565C0', gridcolor='#E3F2FD'),
                              coloraxis_showscale=False,
                              margin=dict(t=10,b=10,l=10,r=110), height=380)
        st.plotly_chart(fig_bar, use_container_width=True)

    if len(display_rows) > 1:
        st.markdown('<div class="section-title">🏢 Account-wise Comparison</div>', unsafe_allow_html=True)
        acct_col1, acct_col2 = st.columns([3,2])
        with acct_col1:
            st.markdown("#### 💳 Account-wise Total Bill")
            fig_acct = px.bar(df_accounts.sort_values('Total (USD)', ascending=False),
                              x='Account Name', y='Total (USD)', color='Total (USD)',
                              color_continuous_scale=[[0,'#87CEEB'],[0.5,'#29B6F6'],[1,'#FF9900']],
                              text='Total (USD)')
            fig_acct.update_traces(texttemplate='$%{text:,.0f}', textposition='outside',
                                   textfont=dict(color='#0D2B4E'),
                                   marker=dict(line=dict(color='#ffffff', width=0.5)))
            fig_acct.update_layout(paper_bgcolor='#ffffff', plot_bgcolor='#ffffff',
                                   font=dict(color='#0D2B4E'), coloraxis_showscale=False,
                                   xaxis=dict(tickangle=-35, color='#1565C0', gridcolor='#E3F2FD'),
                                   yaxis=dict(color='#1565C0', gridcolor='#E3F2FD'),
                                   margin=dict(t=10,b=80), height=340)
            st.plotly_chart(fig_acct, use_container_width=True)
        with acct_col2:
            st.markdown("#### 🍩 Account Cost Share")
            fig_acct_pie = px.pie(df_accounts, names='Account Name', values='Total (USD)',
                                  color_discrete_sequence=['#FF9900','#29B6F6','#26C6DA','#EF5350','#7E57C2','#42A5F5'],
                                  hole=0.5)
            fig_acct_pie.update_traces(textposition='inside', textinfo='percent+label',
                                       textfont=dict(color='white', size=11),
                                       marker=dict(line=dict(color='#ffffff', width=2)))
            fig_acct_pie.update_layout(paper_bgcolor='#ffffff', font=dict(color='#0D2B4E'),
                                       margin=dict(t=10,b=10,l=10,r=10), height=340, showlegend=False)
            st.plotly_chart(fig_acct_pie, use_container_width=True)

    # Margin table
    st.markdown('<div class="section-title">💡 Margin Opportunity</div>', unsafe_allow_html=True)
    st.markdown("""<div class="insight-box">
    <b>How to use this table:</b> Services with the <b>highest cost share</b> have strong demand —
    apply a higher margin there. Services used across <b>multiple accounts</b> are "sticky" —
    clients depend on them, making them ideal for a strategic pricing uplift.
    </div>""", unsafe_allow_html=True)

    svc_account_count = {}
    for acc_data in display_svcs.values():
        for k, v in acc_data.items():
            if k != '_name' and isinstance(v,(int,float)) and v > 0:
                svc_account_count[k] = svc_account_count.get(k,0) + 1

    margin_df = df_svc_totals.copy()
    margin_df['Accounts Using'] = margin_df['Service'].map(lambda s: svc_account_count.get(s,0))
    margin_df['Avg Cost/Account (USD)'] = (margin_df['Total Cost (USD)'] / margin_df['Accounts Using']).round(2)
    total_cost_sum = margin_df['Total Cost (USD)'].sum()
    margin_df['Cost Share (%)'] = (margin_df['Total Cost (USD)'] / total_cost_sum * 100).round(1)

    def margin_suggestion(row):
        if row['Cost Share (%)'] >= 20: return "🔴 High — Apply higher margin"
        elif row['Cost Share (%)'] >= 8: return "🟠 Medium — Moderate margin"
        elif row['Accounts Using'] >= 3: return "🟡 Sticky — High usage, small uplift"
        else: return "🟢 Low usage"

    margin_df['Margin Suggestion'] = margin_df.apply(margin_suggestion, axis=1)
    margin_df = margin_df.rename(columns={'Total Cost (USD)': 'Total Cost (USD) ↓'})
    st.dataframe(
        margin_df[['Service','Total Cost (USD) ↓','Cost Share (%)','Accounts Using',
                   'Avg Cost/Account (USD)','Margin Suggestion']],
        use_container_width=True, hide_index=True,
        column_config={
            'Total Cost (USD) ↓': st.column_config.NumberColumn(format="$%.2f"),
            'Avg Cost/Account (USD)': st.column_config.NumberColumn(format="$%.2f"),
            'Cost Share (%)': st.column_config.ProgressColumn(min_value=0, max_value=100, format="%.1f%%"),
        }
    )

    # Raw tables
    st.markdown('<div class="section-title">📋 Raw Data Tables</div>', unsafe_allow_html=True)
    tab1, tab2 = st.tabs(["🏢 Account Details", "🔧 Service Breakdown"])
    with tab1:
        st.dataframe(df_accounts, use_container_width=True, hide_index=True,
                     column_config={c: st.column_config.NumberColumn(format="$%.2f")
                                    for c in df_accounts.columns if '(USD)' in c})
    with tab2:
        st.dataframe(df_services, use_container_width=True, hide_index=True)

    # Download
    st.divider()
    st.markdown("### 📥 Download Excel Report")
    if selected_payer_key and selected_payer_info:
        dl_label = f"⬇️ Download Report — {selected_payer_info['name']}.xlsx"
        dl_fname = f"AWS_Report_{selected_payer_info['account_id']}.xlsx"
    else:
        dl_label = "⬇️ Download AWS_Invoice_Report.xlsx (All Payers)"
        dl_fname = "AWS_Invoice_Report_All.xlsx"

    excel_cache_key = f"m1_excel_{selected_payer_key or 'all'}"
    if excel_cache_key not in st.session_state:
        with st.spinner("Building Excel report..."):
            st.session_state[excel_cache_key] = build_invoice_excel(display_rows, display_svcs)

    st.download_button(label=dl_label,
                       data=st.session_state[excel_cache_key],
                       file_name=dl_fname,
                       mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                       type="primary", use_container_width=True)


# ═══════════════════════════════════════════════════════════════════════════════
# ══════════════════  MODE 2 — SERVICE-WISE COST BREAKDOWN  ════════════════════
# ═══════════════════════════════════════════════════════════════════════════════
else:
    st.markdown('<div class="section-title">📂 Upload Invoice PDF — Service-wise Breakdown</div>',
                unsafe_allow_html=True)

    st.markdown("""
    <div style="background:#FFF8EC;border-radius:10px;padding:10px 14px;
                border-left:4px solid #FF9900;margin-bottom:12px;">
        <b style="color:#E68A00;">📊 Mode 2 — Service-wise Cost per Account</b><br>
        <span style="color:#0D2B4E;font-size:12px;">
            Upload the same AWS Consolidated Invoice / Statement PDF as Mode 1.<br>
            For every account, it reads <b>Summary for Linked Account</b> and
            <b>Detail for Linked Account</b>, and builds a table of
            <b>Account &times; Service = Net Cost (Charges − Savings Plan)</b>.<br>
            If a service has no Savings Plan, Net Cost is simply its Charges.
        </span>
    </div>
    """, unsafe_allow_html=True)

    bill_files = st.file_uploader(
        "Upload Consolidated Invoice PDF(s)",
        type=["pdf"],
        accept_multiple_files=True,
        key="mode2_uploader",
        help="AWS Consolidated Invoice — contains Summary/Detail for Linked Accounts"
    )

    has_bill = bool(bill_files)

    if has_bill:
        st.success(f"✅ {len(bill_files)} Invoice PDF(s) ready to process")
    else:
        st.info("⬆️ Upload your Invoice PDF(s) and click Submit to begin.")

    uploaded_file_key_m2 = tuple((f.name, getattr(f,"size",0)) for f in (bill_files or []))
    if st.session_state.get("m2_file_key") != uploaded_file_key_m2:
        st.session_state["m2_file_key"] = uploaded_file_key_m2
        st.session_state.pop("m2_parsed_data", None)
        st.session_state.pop("m2_excel", None)
        st.session_state.pop("m2_excel_fname", None)

    submit_m2 = st.button("🚀 Submit & Process Invoice PDF",
                          type="primary", use_container_width=True,
                          disabled=not has_bill)

    if not has_bill:
        st.stop()

    if not submit_m2 and "m2_parsed_data" not in st.session_state:
        st.info("PDF selected. Click 'Submit' to extract and visualize data.")
        st.stop()

    # ── Parse (re-uses the proven Mode-1 parser) ──
    if submit_m2:
        service_data = {}
        errors = []
        progress = st.progress(0, text="Reading Invoice PDF(s)...")
        for i, f in enumerate(bill_files):
            try:
                pdf_bytes = f.getvalue()
                pdf_hash  = _pdf_hash(pdf_bytes)
                _rows, svcs, _p_acct, _p_name = parse_invoice_pdf(pdf_hash, pdf_bytes)
                for acc_id, acc_data in svcs.items():
                    if acc_id not in service_data:
                        service_data[acc_id] = acc_data.copy()
                    else:
                        for k, v in acc_data.items():
                            if k == '_name':
                                service_data[acc_id]['_name'] = v
                            else:
                                service_data[acc_id][k] = round(service_data[acc_id].get(k,0.0)+v, 2)
            except Exception as e:
                errors.append(f"{f.name}: {e}")
            progress.progress(int((i+1)/len(bill_files)*100), text=f"{f.name} done")
        progress.empty()
        for e in errors:
            st.error(f"❌ {e}")
        st.session_state["m2_parsed_data"] = {"service_data": service_data}

    data = st.session_state.get("m2_parsed_data", {})
    service_data = data.get("service_data", {})

    if not service_data:
        st.warning("⚠️ No service data found. Please verify you have uploaded a valid Consolidated Invoice PDF.")
        st.stop()

    df_pivot = build_service_pivot_df(service_data)
    service_cols = [c for c in df_pivot.columns if c != 'Account ID']

    # KPI Cards
    st.markdown('<div class="section-title">📊 Summary</div>', unsafe_allow_html=True)
    grand_total  = df_pivot[service_cols].sum().sum()
    total_accts  = len(df_pivot)
    unique_svcs  = len(service_cols)
    svc_totals_s = df_pivot[service_cols].sum().sort_values(ascending=False)
    top_svc      = svc_totals_s.index[0] if len(svc_totals_s) else "N/A"
    top_svc_cost = svc_totals_s.iloc[0] if len(svc_totals_s) else 0.0

    c1, c2, c3, c4 = st.columns(4)
    def kpi2(col, icon, label, value, sub=""):
        col.markdown(f"""
        <div class="metric-card">
          <div class="metric-label">{icon} {label}</div>
          <div class="metric-value">{value}</div>
          <div class="metric-sub">{sub}</div>
        </div>""", unsafe_allow_html=True)

    kpi2(c1, "💰", "Grand Total",  f"${grand_total:,.2f}", "Charges - Savings Plan")
    kpi2(c2, "🏢", "Accounts",     str(total_accts), "accounts")
    kpi2(c3, "🔧", "Services",     str(unique_svcs), "unique services")
    kpi2(c4, "🔝", "Top Service",
         top_svc[:16]+"…" if len(top_svc)>16 else top_svc, f"${top_svc_cost:,.2f}")

    st.markdown("<br>", unsafe_allow_html=True)

    # ── Charts ──
    svc_summary = svc_totals_s.reset_index()
    svc_summary.columns = ['Service', 'Cost (USD)']

    chart1, chart2 = st.columns([1,1])

    with chart1:
        st.markdown("#### 🥧 Service Cost Distribution")
        fig_pie2 = px.pie(svc_summary.head(10), names='Service', values='Cost (USD)',
                          color_discrete_sequence=px.colors.sequential.Oranges_r, hole=0.35)
        fig_pie2.update_traces(textposition='inside', textinfo='percent+label',
                               textfont=dict(color='white', size=10),
                               marker=dict(line=dict(color='#ffffff', width=1.5)))
        fig_pie2.update_layout(paper_bgcolor='#ffffff', font=dict(color='#0D2B4E'),
                               margin=dict(t=10,b=10,l=10,r=10), height=380)
        st.plotly_chart(fig_pie2, use_container_width=True)

    with chart2:
        st.markdown("#### 📊 Top Services by Cost")
        fig_bar2 = px.bar(svc_summary.head(12), x='Cost (USD)', y='Service', orientation='h',
                          color='Cost (USD)',
                          color_continuous_scale=[[0,'#FFE0B2'],[0.5,'#FF9900'],[1,'#E65100']],
                          text='Cost (USD)')
        fig_bar2.update_traces(texttemplate='$%{text:,.0f}', textposition='outside',
                               textfont=dict(color='#0D2B4E', size=11))
        fig_bar2.update_layout(paper_bgcolor='#ffffff', plot_bgcolor='#ffffff',
                               font=dict(color='#0D2B4E'),
                               yaxis=dict(autorange="reversed", color='#1565C0', gridcolor='#E3F2FD'),
                               xaxis=dict(color='#1565C0', gridcolor='#E3F2FD'),
                               coloraxis_showscale=False,
                               margin=dict(t=10,b=10,l=10,r=110), height=380)
        st.plotly_chart(fig_bar2, use_container_width=True)

    # Account-wise total chart (sum across all services)
    acct_totals = df_pivot.copy()
    acct_totals['Total (USD)'] = acct_totals[service_cols].sum(axis=1)
    acct_totals = acct_totals.sort_values('Total (USD)', ascending=False)
    st.markdown("#### 🏢 Account-wise Total (Net Cost)")
    fig_acct = px.bar(acct_totals, x='Account ID', y='Total (USD)', color='Total (USD)',
                     color_continuous_scale=[[0,'#87CEEB'],[0.5,'#29B6F6'],[1,'#0D47A1']],
                     text='Total (USD)')
    fig_acct.update_traces(texttemplate='$%{text:,.0f}', textposition='outside',
                          textfont=dict(color='#0D2B4E'))
    fig_acct.update_layout(paper_bgcolor='#ffffff', plot_bgcolor='#ffffff',
                          font=dict(color='#0D2B4E'), coloraxis_showscale=False,
                          xaxis=dict(tickangle=-30, color='#1565C0', gridcolor='#E3F2FD'),
                          yaxis=dict(color='#1565C0', gridcolor='#E3F2FD'),
                          margin=dict(t=10,b=80), height=340)
    st.plotly_chart(fig_acct, use_container_width=True)

    # Filter & Table (Account x Service pivot)
    st.markdown('<div class="section-title">🔍 Account &times; Service Table</div>', unsafe_allow_html=True)
    svc_filter = st.multiselect("Service filter (columns to show):",
                                options=service_cols, default=[])

    display_cols = ['Account ID'] + (svc_filter if svc_filter else service_cols)
    filtered_df = df_pivot[display_cols].copy()

    st.caption(f"Showing {len(filtered_df):,} accounts | Grand Total: ${filtered_df[[c for c in display_cols if c!='Account ID']].sum().sum():,.2f}")
    st.dataframe(filtered_df, use_container_width=True, hide_index=True,
                 column_config={c: st.column_config.NumberColumn(format="$%.2f")
                                for c in display_cols if c != 'Account ID'})

    # Download
    st.divider()
    st.markdown("### 📥 Download Excel Report")

    if "m2_excel" not in st.session_state:
        with st.spinner("Building Excel report..."):
            st.session_state["m2_excel"] = build_service_pivot_excel(df_pivot)

    st.download_button(
        label="⬇️ Download AWS_ServiceBreakdown.xlsx",
        data=st.session_state["m2_excel"],
        file_name="AWS_ServiceBreakdown.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type="primary", use_container_width=True,
    )


# ── Footer ──
st.markdown(
    "<div style='text-align:center;color:#1565C0;font-size:12px;margin-top:16px;padding:12px;"
    "border-top:1px solid #90CAF9;'>"
    "☁️ &nbsp;AWS Invoice Dashboard &nbsp;•&nbsp; Built with Streamlit + Plotly &nbsp;•&nbsp; "
    "<span style='color:#FF9900;'>Powered by AWS Analytics</span>"
    "</div>",
    unsafe_allow_html=True
)
